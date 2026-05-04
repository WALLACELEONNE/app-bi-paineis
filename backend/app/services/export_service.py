import io
from datetime import date
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from app.models import Transacao
from app.services.bi_service import _filter_transacoes
import openpyxl
import structlog

logger = structlog.get_logger()


async def exportar_excel(
    db: AsyncSession,
    tenant_id: int,
    params: dict,
    filename: str,
) -> bytes:
    query = select(Transacao)
    query = _filter_transacoes(query, tenant_id, params)
    query = query.order_by(Transacao.data_competencia.desc()).limit(
        params.get("limit", 5000)
    )

    result = await db.execute(query)
    rows = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"

    headers = [
        "ID",
        "Data Competência",
        "Departamento",
        "Valor",
        "Quantidade",
        "Unidade",
        "Centro Custo",
        "Produto",
        "Fazenda",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    for i, trans in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=trans.id)
        ws.cell(
            row=i,
            column=2,
            value=str(trans.data_competencia) if trans.data_competencia else "",
        )
        ws.cell(row=i, column=3, value=trans.departamento_id)
        ws.cell(row=i, column=4, value=float(trans.valor) if trans.valor else 0)
        ws.cell(
            row=i, column=5, value=float(trans.quantidade) if trans.quantidade else None
        )
        ws.cell(row=i, column=6, value=trans.unidade)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("export_excel", filename=filename, rows=len(rows))
    return output.getvalue()
